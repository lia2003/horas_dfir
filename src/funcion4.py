"""
Función 4 — Resumen de horas YA cargadas (Excel plano por persona-engagement).

NOTA: el reporte que se envía al jefe de RD ya NO se genera acá — se genera
en la Función 3 (Reporte_RD.xlsx), a partir de las horas aprobadas para el
equipo, no de lo que ya quedó cargado en el Excel. Esta función queda solo
como resumen interno/posterior de verificación, una vez que el equipo
efectivamente cargó sus horas.

Filtra filas con Estado='Cargado' y Tipo='Cliente' (incluye 1. Lunes (FDS)).
Si Andrea Neira tiene horas, pregunta su prorateo y lo aplica.
Guarda el resultado en Resumen_Horas_Cargadas.xlsx en la carpeta de salida
de la semana.

Columnas del Excel de salida:
  A: NSR Rate       — multiplicador de Rates FY26 según el rank de la persona
  B: Person Name    — "Rank - Nombre completo"
  C: Engagement Number
  D: Hours
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .excel_reader import (
    PROYECTOS_EXCLUIDOS,
    get_semana_rows,
    load_jobs,
    load_rates,
    parse_semana_fecha,
)


def generar_resumen(excel_path: Path, semana_nombre: str, output_dir: Path) -> None:
    parse_semana_fecha(semana_nombre)

    filas = get_semana_rows(excel_path, semana_nombre)
    jobs  = load_jobs(excel_path)
    rates = load_rates(excel_path)

    # ── 1. Filtrar: Cargado + Cliente + no excluidos ──────────────────────────
    filas_cargadas = [
        f for f in filas
        if f["estado"] == "Cargado"
        and f["tipo"] == "Cliente"
        and (f["proyecto"] or "").strip() not in PROYECTOS_EXCLUIDOS
        and f["proyecto"]
        and (f["horas"] or 0) > 0
    ]

    if not filas_cargadas:
        print("  No hay filas en Estado='Cargado' para generar el resumen.")
        return

    # ── 2. Detectar personas con prorateo y pedirlo ───────────────────────────
    nombres_en_datos = {f["nombre"] for f in filas_cargadas if f["nombre"]}

    nombre_andrea = next(
        (n for n in nombres_en_datos
         if n and "andrea" in n.lower() and "neira" in n.lower()),
        None,
    )
    prorateo_andrea: float | None = None
    if nombre_andrea:
        print(f"\n  Se detectaron horas cargadas para {nombre_andrea}.")
        while True:
            raw = input(
                f"  Prorateo de {nombre_andrea} "
                "(ej: 0.48 — sus horas se multiplicarán): "
            ).strip()
            try:
                prorateo_andrea = float(raw.replace(",", "."))
                break
            except ValueError:
                print("  Ingresa un número válido (ej: 0.48).")

    nombre_daniel = next(
        (n for n in nombres_en_datos
         if n and "daniel" in n.lower() and "cabrera" in n.lower()),
        None,
    )
    prorateo_daniel: float | None = None
    if nombre_daniel:
        print(f"\n  Se detectaron horas cargadas para {nombre_daniel}.")
        while True:
            raw = input(
                f"  Prorateo de {nombre_daniel} "
                "(ej: 0.48 — sus horas se multiplicarán): "
            ).strip()
            try:
                prorateo_daniel = float(raw.replace(",", "."))
                break
            except ValueError:
                print("  Ingresa un número válido (ej: 0.48).")

    RATE_DANIEL = 338

    # ── 3. Acumular horas por (nombre, rank, engagement) ─────────────────────
    # Clave: (nombre, rank, engagement) → horas totales originales
    datos: dict[tuple[str, str, str], float] = defaultdict(float)
    sin_engagement: list[dict] = []

    for f in filas_cargadas:
        nombre = f["nombre"] or ""
        rank   = f["rank"]   or ""
        proy   = (f["proyecto"] or "").strip()
        info   = jobs.get(proy, {})
        eng    = (info.get("engagement") or "").strip()

        if not eng:
            sin_engagement.append(f)
            eng = f"SIN ENGAGEMENT ({proy})"

        datos[(nombre, rank, eng)] += f["horas"]

    # ── 4. Avisar proyectos sin engagement ────────────────────────────────────
    if sin_engagement:
        print("\n  ADVERTENCIA — Las siguientes filas cargadas no tienen engagement "
              "en Jobs FY26:")
        proyectos_vistos: set[str] = set()
        for f in sin_engagement:
            proy = (f["proyecto"] or "").strip()
            if proy not in proyectos_vistos:
                print(f"    - {proy} (gerente: {jobs.get(proy, {}).get('gerente', 'N/A')})")
                proyectos_vistos.add(proy)
        print("  Actualiza el engagement en Jobs FY26 y vuelve a generar el resumen.\n")

    # ── 5. Generar Excel ──────────────────────────────────────────────────────
    carpeta = output_dir / semana_nombre
    carpeta.mkdir(parents=True, exist_ok=True)
    archivo = carpeta / "Resumen_Horas_Cargadas.xlsx"

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Resumen"

    # Estilos
    hdr_font   = Font(bold=True, color="FFFFFF")
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    center_aln = Alignment(horizontal="center", vertical="center")
    left_aln   = Alignment(horizontal="left",   vertical="center")
    right_aln  = Alignment(horizontal="right",  vertical="center")

    # Encabezado
    headers = ["NSR Rate", "Person Name", "Engagement Number", "Hours"]
    for col_idx, titulo in enumerate(headers, start=1):
        cell = ws_out.cell(row=1, column=col_idx, value=titulo)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center_aln

    ws_out.column_dimensions["A"].width = 12
    ws_out.column_dimensions["B"].width = 38
    ws_out.column_dimensions["C"].width = 22
    ws_out.column_dimensions["D"].width = 10

    def _nsr_val(nombre: str, rank: str) -> float:
        v = RATE_DANIEL if (nombre_daniel and nombre == nombre_daniel) else rates.get(rank, 0)
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    claves_ordenadas = sorted(
        datos.keys(),
        key=lambda x: -_nsr_val(x[0], x[1]),
    )

    # Filas de datos — ordenadas de mayor a menor NSR Rate
    fila_actual = 2
    for (nombre, rank, eng) in claves_ordenadas:
        h_orig = datos[(nombre, rank, eng)]

        if nombre_andrea and nombre == nombre_andrea and prorateo_andrea is not None:
            horas = round(h_orig * prorateo_andrea, 1)
        elif nombre_daniel and nombre == nombre_daniel and prorateo_daniel is not None:
            horas = round(h_orig * prorateo_daniel, 1)
        else:
            horas = h_orig

        if nombre_daniel and nombre == nombre_daniel:
            nsr_rate = RATE_DANIEL
        else:
            nsr_rate = rates.get(rank, "")
        person_name = f"{rank} - {nombre}" if rank else nombre

        ws_out.cell(row=fila_actual, column=1, value=nsr_rate).alignment = center_aln
        ws_out.cell(row=fila_actual, column=2, value=person_name).alignment = left_aln
        ws_out.cell(row=fila_actual, column=3, value=eng).alignment = left_aln
        c = ws_out.cell(row=fila_actual, column=4, value=horas)
        c.alignment    = right_aln
        c.number_format = "0.0"
        fila_actual += 1

    wb_out.save(archivo)

    # ── 6. Mostrar resumen en pantalla ────────────────────────────────────────
    print(f"\n  === RESUMEN SEMANAL - {semana_nombre} ===\n")
    for (nombre, rank, eng) in claves_ordenadas:
        h_orig = datos[(nombre, rank, eng)]
        if nombre_andrea and nombre == nombre_andrea and prorateo_andrea is not None:
            h = round(h_orig * prorateo_andrea, 1)
            label_h = f"{h:.1f}h  (×{prorateo_andrea})"
        elif nombre_daniel and nombre == nombre_daniel and prorateo_daniel is not None:
            h = round(h_orig * prorateo_daniel, 1)
            label_h = f"{h:.1f}h  (×{prorateo_daniel})"
        else:
            h = h_orig
            label_h = f"{h:.1f}h"
        if nombre_daniel and nombre == nombre_daniel:
            nsr = RATE_DANIEL
        else:
            nsr = rates.get(rank, "?")
        print(f"  [{nsr}]  {rank} - {nombre:<28} {eng:<20} {label_h}")

    print(f"\n  Guardado en: {archivo}")
