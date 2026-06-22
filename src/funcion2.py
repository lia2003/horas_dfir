"""
Función 2 — Arrastre de horas a la semana siguiente (FDS).

Flujo:
  1. Muestra filas Cliente con Estado=Pendiente agrupadas por proyecto.
  2. El usuario elige cuáles arrastrar.
  3. Si la hoja siguiente no existe, la crea clonando 'Plantilla' con todos
     sus atributos: Table (filtros), validaciones de datos (desplegables) y
     formato condicional (colores).
  4. Inserta las filas seleccionadas como '1. Lunes (FDS)' / Estado='Pendiente'
     ANTES del primer día no-FDS de cada persona, manteniendo el orden.
"""

from __future__ import annotations

import copy
import os
import re
import zipfile
from collections import defaultdict
from datetime import timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment

from .excel_reader import (
    PROYECTOS_EXCLUIDOS,
    get_semana_rows,
    parse_semana_fecha,
)
from .utils import guardar_excel, confirmar

DIA_FDS = "1. Lunes (FDS)"


def arrastrar_fds(excel_path: Path, semana_nombre: str) -> None:
    lunes_actual     = parse_semana_fecha(semana_nombre)
    lunes_siguiente  = lunes_actual + timedelta(weeks=1)
    nombre_siguiente = f"Semana {lunes_siguiente.strftime('%d-%m-%Y')}"

    # ── 1. Leer filas pendientes Cliente ──────────────────────────────────────
    filas = get_semana_rows(excel_path, semana_nombre)
    pendientes = [
        f for f in filas
        if f["tipo"] == "Cliente"
        and f["estado"] == "Pendiente"
        and (f["proyecto"] or "").strip() not in PROYECTOS_EXCLUIDOS
        and f["proyecto"]
    ]

    if not pendientes:
        print("  No hay filas Cliente con Estado=Pendiente en esta semana.")
        _ofrecer_crear_hoja(excel_path, nombre_siguiente)
        return

    # ── 2. Mostrar agrupadas por proyecto ─────────────────────────────────────
    por_proyecto: dict[str, list[dict]] = defaultdict(list)
    for f in pendientes:
        por_proyecto[(f["proyecto"] or "").strip()].append(f)

    print(f"\n  Filas Pendiente por proyecto  (destino: {nombre_siguiente}):\n")
    numero_a_fila: dict[int, dict] = {}
    num = 1
    for proy, fps in sorted(por_proyecto.items()):
        print(f"  Proyecto: {proy}")
        for f in fps:
            tarea_display = (f["tarea"] or "(sin tarea)").strip()[:55]
            print(f"    [{num:>2}] {f['nombre']:<25} | {f['dia']:<18} | "
                  f"{f['horas']:>5}h | {tarea_display}")
            numero_a_fila[num] = f
            num += 1
        print()

    # ── 3. Selección del usuario ──────────────────────────────────────────────
    seleccionadas = _pedir_seleccion(numero_a_fila)
    if seleccionadas is None:
        print("  Operación cancelada.")
        return

    print(f"\n  Seleccionadas para arrastrar: {len(seleccionadas)}")
    for f in seleccionadas:
        print(f"    - {f['nombre']:<25} | {f['proyecto']:<20} | {f['horas']}h")

    if not confirmar(f"\n  ¿Confirmar escritura en '{nombre_siguiente}'?"):
        print("  Operación cancelada.")
        return

    # ── 4. Escribir en el Excel ───────────────────────────────────────────────
    _escribir_fds(excel_path, nombre_siguiente, seleccionadas, semana_nombre)
    print(f"\n  Listo. {len(seleccionadas)} fila(s) copiadas a '{nombre_siguiente}' "
          f"como '{DIA_FDS}'.")


# ── helpers de selección ──────────────────────────────────────────────────────

def _pedir_seleccion(numero_a_fila: dict[int, dict]) -> list[dict] | None:
    """Retorna la lista de filas seleccionadas, o None si el usuario cancela."""
    print("  Ingresa los números a arrastrar separados por coma  "
          "(ej: 1,3,5),")
    print("  'todos' para seleccionar todos, o 'ninguno' para cancelar:")
    while True:
        resp = input("  > ").strip().lower()
        if resp == "ninguno":
            return None
        if resp == "todos":
            return list(numero_a_fila.values())
        try:
            nums = [int(x.strip()) for x in resp.split(",") if x.strip()]
            invalidos = [n for n in nums if n not in numero_a_fila]
            if invalidos:
                print(f"  Números inválidos: {invalidos}. Intenta de nuevo.")
                continue
            seleccionadas = [numero_a_fila[n] for n in nums]
            if not seleccionadas:
                print("  Debes seleccionar al menos una fila.")
                continue
            return seleccionadas
        except ValueError:
            print("  Formato inválido. Usa números separados por coma.")


# ── escritura en el Excel ─────────────────────────────────────────────────────

def _escribir_fds(
    excel_path: Path,
    nombre_siguiente: str,
    filas_fds: list[dict],
    semana_nombre: str,
) -> None:
    # Leer el bloque x14 de la Plantilla ANTES de abrir con openpyxl
    ext_lst = _leer_ext_lst_plantilla(excel_path)

    wb = openpyxl.load_workbook(excel_path)
    hoja_nueva = nombre_siguiente not in wb.sheetnames

    if hoja_nueva:
        _clonar_plantilla(wb, nombre_siguiente)
        print(f"  Hoja '{nombre_siguiente}' creada desde Plantilla.")
    else:
        print(f"  Hoja '{nombre_siguiente}' ya existe; se insertarán filas en orden.")

    ws = wb[nombre_siguiente]
    # Hoja origen para copiar estilos de celda (mismas columnas, misma fila real)
    ws_src = wb[semana_nombre]

    # Agrupar las FDS por nombre de persona (manteniendo orden de aparición)
    fds_por_persona: dict[str, list[dict]] = defaultdict(list)
    orden_personas: list[str] = []
    for f in filas_fds:
        nombre = f["nombre"] or ""
        if nombre not in fds_por_persona:
            orden_personas.append(nombre)
        fds_por_persona[nombre].append(f)

    for nombre in orden_personas:
        nuevas = fds_por_persona[nombre]
        fila_insercion = _buscar_primera_fila_no_fds(ws, nombre)

        if fila_insercion is None:
            for f in nuevas:
                row_idx = ws.max_row + 1
                _escribir_celda_fds(ws, row_idx, f, ws_src)
        else:
            n = len(nuevas)
            ws.insert_rows(fila_insercion, n)
            for i, f in enumerate(nuevas):
                _escribir_celda_fds(ws, fila_insercion + i, f, ws_src)

    # Post-proceso: inyectar x14:dataValidations y restaurar archivos vulnerables
    def _post(tmp):
        if hoja_nueva:
            _inyectar_ext_lst(tmp, nombre_siguiente, ext_lst, excel_path)

    guardar_excel(wb, excel_path, post_process=_post)
    wb.close()


# ── validaciones x14 (extensión que openpyxl no soporta) ─────────────────────

def _leer_ext_lst_plantilla(excel_path: Path) -> str:
    """
    Extrae el bloque <extLst> de la hoja Plantilla directamente del zip xlsx.
    Contiene las x14:dataValidations (Rank, Tipo, Tipo actividad, Día, Estado)
    que openpyxl ignora al leer y al escribir.
    """
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            wb_xml = z.read('xl/workbook.xml').decode('utf-8')
            # Dos pasos: primero el elemento <sheet>, luego extraer r:id
            # (el orden de atributos varía entre versiones de Excel)
            sheet_elem = re.search(r'<sheet\b[^>]*\bname="Plantilla"[^>]*>', wb_xml)
            if not sheet_elem:
                return ''
            rid_m = re.search(r'\br:id="(rId\d+)"', sheet_elem.group(0))
            if not rid_m:
                return ''
            rid = rid_m.group(1)

            rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            rel_m = re.search(
                rf'<Relationship\b(?=[^>]*\bId="{re.escape(rid)}")[^>]*/>', rels
            )
            if not rel_m:
                return ''
            tm = re.search(r'\bTarget="([^"]+)"', rel_m.group(0))
            if not tm:
                return ''

            target = tm.group(1)
            sheet_key = target.lstrip('/') if target.startswith('/') else f'xl/{target}'
            sheet_xml = z.read(sheet_key).decode('utf-8')
            m3 = re.search(r'<extLst>.*?</extLst>', sheet_xml, re.DOTALL)
            return m3.group(0) if m3 else ''
    except Exception:
        return ''


def _inyectar_ext_lst(
    tmp_path: Path,
    nuevo_nombre: str,
    ext_lst: str,
    original_path: Path | None = None,
) -> None:
    """
    Abre el xlsx temporal como zip, localiza la hoja recién creada e inyecta
    el bloque extLst (x14:dataValidations) antes de </worksheet>.
    Elimina los xr:uid para evitar GUIDs duplicados entre hojas.

    También restaura externalLinks y drawings desde el archivo original porque
    openpyxl puede corromper sus cached values al guardar, lo que provoca el
    diálogo "We found a problem…" al abrir el Excel.
    """
    if not ext_lst:
        return

    ext_lst_limpio = re.sub(r'\s+xr:uid="\{[^}]+\}"', '', ext_lst)

    try:
        with zipfile.ZipFile(tmp_path, 'r') as z:
            wb_xml = z.read('xl/workbook.xml').decode('utf-8')
            # Dos pasos: primero el elemento <sheet>, luego extraer r:id
            sheet_elem = re.search(
                rf'<sheet\b[^>]*\bname="{re.escape(nuevo_nombre)}"[^>]*>', wb_xml
            )
            if not sheet_elem:
                return
            rid_m = re.search(r'\br:id="(rId\d+)"', sheet_elem.group(0))
            if not rid_m:
                return
            rid = rid_m.group(1)

            rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            rel_m = re.search(
                rf'<Relationship\b(?=[^>]*\bId="{re.escape(rid)}")[^>]*/>', rels
            )
            if not rel_m:
                return
            tm = re.search(r'\bTarget="([^"]+)"', rel_m.group(0))
            if not tm:
                return

            target = tm.group(1)
            sheet_key = target.lstrip('/') if target.startswith('/') else f'xl/{target}'

            # Preservar ZipInfo de cada entrada (compresión, timestamps, etc.)
            all_infos = {info.filename: info for info in z.infolist()}
            all_entries = {name: z.read(name) for name in z.namelist()}

        # Modificar solo la hoja nueva: siempre reemplazar extLst completo
        # (copy_worksheet puede haber copiado uno incompleto que bloquea la inyección)
        sheet_xml = all_entries[sheet_key].decode('utf-8')
        sheet_xml = re.sub(r'<extLst>.*?</extLst>', '', sheet_xml, flags=re.DOTALL)
        sheet_xml = sheet_xml.replace('</worksheet>', f'{ext_lst_limpio}</worksheet>', 1)
        all_entries[sheet_key] = sheet_xml.encode('utf-8')

        # Restaurar externalLinks y drawings desde el archivo original:
        # openpyxl corrompe sus cached values al hacer wb.save(), lo que genera
        # el diálogo de reparación "We found a problem with some content".
        # Nuestros cambios (nueva hoja semanal) no modifican estos archivos,
        # así que es seguro restaurarlos tal cual estaban.
        if original_path and original_path.exists():
            try:
                with zipfile.ZipFile(original_path, 'r') as z_orig:
                    orig_infos = {info.filename: info for info in z_orig.infolist()}
                    for name in z_orig.namelist():
                        if 'externalLinks' in name or 'drawings' in name:
                            all_entries[name] = z_orig.read(name)
                            all_infos[name] = orig_infos[name]
            except Exception:
                pass  # Si falla la restauración, continuar con lo que hay

        # Reescribir el ZIP preservando el ZipInfo original de cada entrada
        tmp2 = tmp_path.with_suffix('.tmp2')
        with zipfile.ZipFile(tmp2, 'w', allowZip64=True) as z_out:
            for name, data in all_entries.items():
                if name in all_infos:
                    z_out.writestr(all_infos[name], data)
                else:
                    z_out.writestr(name, data)
        os.replace(tmp2, tmp_path)

    except Exception as e:
        print(f"  AVISO: No se pudieron inyectar validaciones extendidas: {e}")


# ── clonación de Plantilla ────────────────────────────────────────────────────

def _clonar_plantilla(wb, nuevo_nombre: str) -> None:
    """
    Clona la hoja 'Plantilla' preservando Table (filtros), validaciones de
    datos (desplegables) y formato condicional (colores por Tipo y Estado).
    openpyxl.copy_worksheet no copia estos atributos, por eso se agregan
    manualmente después de la copia básica.
    """
    if "Plantilla" not in wb.sheetnames:
        raise KeyError("No se encontró la hoja 'Plantilla' para clonar.")

    src = wb["Plantilla"]

    # 1. Copia básica: celdas, estilos, propiedades de hoja
    dst = wb.copy_worksheet(src)
    dst.title = nuevo_nombre

    # 2. Dimensiones de columnas y filas (anchos y altos)
    for col, dim in src.column_dimensions.items():
        dst.column_dimensions[col] = copy.copy(dim)
    for row_num, dim in src.row_dimensions.items():
        dst.row_dimensions[row_num] = copy.copy(dim)

    # 3. Validaciones de datos (desplegables en celdas: Tipo, Proyecto, etc.)
    # Limpiar primero lo que copy_worksheet pudo haber copiado parcialmente
    dst.data_validations.dataValidation = []
    for dv in src.data_validations.dataValidation:
        dst.add_data_validation(copy.deepcopy(dv))

    # 4. Formato condicional — limpiar antes de re-agregar para evitar duplicados
    dst.conditional_formatting._cf_rules = {}
    for cf in src.conditional_formatting:
        sqref = str(cf.sqref)
        for rule in cf.cfRule:
            dst.conditional_formatting.add(sqref, copy.deepcopy(rule))

    # 5. Tabla (da los filtros de columna y el estilo banded de la tabla)
    _copiar_tabla(wb, src, dst)

    # 6. Limpiar referencias a charts/imágenes heredadas por copy_worksheet:
    #    si se dejan, openpyxl escribe una relación a drawing1.xml que ya
    #    pertenece a Plantilla → Excel reporta "Drawing shape" al abrir.
    dst._charts = []
    dst._images = []


def _copiar_tabla(wb, src, dst) -> None:
    """Copia la Tabla de src a dst con un nombre e id únicos en el libro."""
    if not src.tables:
        return

    # Calcular máximo id de tabla existente en todo el libro
    max_id = 0
    for ws_name in wb.sheetnames:
        if ws_name == dst.title:
            continue
        ws_tmp = wb[ws_name]
        for t_name in ws_tmp.tables:
            t = ws_tmp.tables[t_name]
            if t.id and t.id > max_id:
                max_id = t.id

    for i, t_name in enumerate(src.tables):
        new_table = copy.deepcopy(src.tables[t_name])
        new_id = max_id + 1 + i
        new_table.id = new_id
        new_table.displayName = f"Tabla{new_id}"
        new_table.name = new_table.displayName
        dst.add_table(new_table)


# ── utilidad: crear hoja aunque no haya FDS ──────────────────────────────────

def _ofrecer_crear_hoja(excel_path: Path, nombre_siguiente: str) -> None:
    """Si no hay pendientes, ofrece crear la hoja de la semana siguiente igual."""
    wb_check = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    existe = nombre_siguiente in wb_check.sheetnames
    wb_check.close()

    if existe:
        print(f"  La hoja '{nombre_siguiente}' ya existe.")
        return

    if confirmar(f"  ¿Crear la hoja '{nombre_siguiente}' desde Plantilla de todas formas?"):
        ext_lst = _leer_ext_lst_plantilla(excel_path)
        wb = openpyxl.load_workbook(excel_path)
        _clonar_plantilla(wb, nombre_siguiente)

        def _post(tmp):
            _inyectar_ext_lst(tmp, nombre_siguiente, ext_lst, excel_path)

        guardar_excel(wb, excel_path, post_process=_post)
        wb.close()
        print(f"  Hoja '{nombre_siguiente}' creada.")


# ── helpers de inserción ──────────────────────────────────────────────────────

def _buscar_primera_fila_no_fds(ws, nombre: str) -> int | None:
    """
    Retorna el número de fila (1-based) de la primera fila de 'nombre'
    cuyo valor en col H NO sea '1. Lunes (FDS)'.
    Retorna None si la persona no está en la hoja.
    """
    for fila in ws.iter_rows(min_row=2):
        nombre_celda = fila[0].value
        dia_celda    = fila[7].value if len(fila) > 7 else None
        if nombre_celda == nombre:
            if dia_celda != DIA_FDS:
                return fila[0].row
    return None


def _fila_a_valores(f: dict) -> list:
    """Convierte un dict de fila en lista para ws.append()."""
    return [
        f["nombre"],
        f["rank"],
        f["tipo"],
        f["proyecto"],
        f["tipo_actividad"],
        f["tarea"],
        f["horas"],
        DIA_FDS,
        "Pendiente",
        None,   # Cargado en Job
        None,   # Comentarios
    ]


def _copiar_estilo(src, dst) -> None:
    """Copia font, fill, border, alignment y number_format de src a dst.

    Construye nuevos objetos de estilo en vez de usar copy.deepcopy porque
    StyleProxy de openpyxl no es compatible con deepcopy (recursión infinita
    en __getattr__ al buscar __setstate__).
    """
    if not src.has_style:
        return
    try:
        f = src.font
        dst.font = Font(
            name=f.name, bold=f.bold, italic=f.italic,
            vertAlign=f.vertAlign, underline=f.underline,
            strike=f.strike, size=f.size, color=f.color,
        )
    except Exception:
        pass
    try:
        fi = src.fill
        dst.fill = PatternFill(
            fill_type=fi.fill_type, fgColor=fi.fgColor, bgColor=fi.bgColor,
        )
    except Exception:
        pass
    try:
        b = src.border
        dst.border = Border(
            left=b.left, right=b.right, top=b.top, bottom=b.bottom,
            diagonal=b.diagonal, diagonal_direction=b.diagonal_direction,
        )
    except Exception:
        pass
    try:
        a = src.alignment
        dst.alignment = Alignment(
            horizontal=a.horizontal, vertical=a.vertical,
            text_rotation=a.text_rotation, wrap_text=a.wrap_text,
            shrink_to_fit=a.shrink_to_fit, indent=a.indent,
        )
    except Exception:
        pass
    try:
        dst.number_format = src.number_format
    except Exception:
        pass


def _escribir_celda_fds(ws, row_num: int, f: dict, ws_src=None) -> None:
    """Escribe los valores FDS copiando el estilo de la fila origen."""
    valores = _fila_a_valores(f)
    src_row = f.get("fila_excel")
    for col_idx, val in enumerate(valores, start=1):
        dst = ws.cell(row=row_num, column=col_idx, value=val)
        if ws_src is not None and src_row:
            _copiar_estilo(ws_src.cell(row=src_row, column=col_idx), dst)
        else:
            _copiar_estilo(ws.cell(row=2, column=col_idx), dst)
