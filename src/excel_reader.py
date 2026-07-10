"""Lectura del Excel: detección de semanas, filas semanales, Jobs FY26, Rates FY26."""

import re
from datetime import date
from pathlib import Path

import openpyxl

# Regex para detectar hojas semanales
SEMANA_RE = re.compile(r"^Semana (\d{2})-(\d{2})-(\d{4})$")

# Proyectos excluidos de todas las funciones (match exacto sobre col D de la hoja semanal)
PROYECTOS_EXCLUIDOS = {"LE", "LE Gestión", "LE Gestion"}

# Índices de columnas en la hoja semanal (0-based, fila 1 = header)
COL_NOMBRE       = 0
COL_RANK         = 1
COL_TIPO         = 2
COL_PROYECTO     = 3
COL_TIPO_ACT     = 4
COL_TAREA        = 5
COL_HORAS        = 6
COL_DIA          = 7
COL_ESTADO       = 8
COL_CARGADO_JOB  = 9
COL_COMENTARIOS  = 10


def detect_semanas(excel_path: Path) -> list[tuple[date, str]]:
    """
    Retorna lista de (date_lunes, nombre_hoja) para todas las hojas semanales,
    ordenada por fecha ascendente.
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    semanas = []
    for nombre in wb.sheetnames:
        m = SEMANA_RE.match(nombre)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                semanas.append((date(y, mo, d), nombre))
            except ValueError:
                pass  # fecha inválida, ignorar
    wb.close()
    return sorted(semanas)


def parse_semana_fecha(semana_nombre: str) -> date:
    """'Semana DD-MM-YYYY' → date del lunes."""
    m = SEMANA_RE.match(semana_nombre)
    if not m:
        raise ValueError(f"Nombre de semana inválido: {semana_nombre!r}")
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return date(y, mo, d)


def get_semana_rows(excel_path: Path, semana_nombre: str) -> list[dict]:
    """
    Lee todas las filas de datos de una hoja semanal.
    Retorna lista de dicts con claves normalizadas. Omite filas completamente vacías.
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if semana_nombre not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Hoja '{semana_nombre}' no existe en el Excel.")
    ws = wb[semana_nombre]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    resultado = []
    for i, fila in enumerate(rows[1:], start=2):  # fila 1 = header, empezamos en 2
        if not any(c is not None for c in fila):
            continue
        resultado.append({
            "fila_excel": i,
            "nombre":        _str(fila, COL_NOMBRE),
            "rank":          _str(fila, COL_RANK),
            "tipo":          _str(fila, COL_TIPO),
            "proyecto":      _str(fila, COL_PROYECTO),
            "tipo_actividad":_str(fila, COL_TIPO_ACT),
            "tarea":         _str(fila, COL_TAREA),
            "horas":         _num(fila, COL_HORAS),
            "dia":           _str(fila, COL_DIA),
            "estado":        _str(fila, COL_ESTADO),
            "cargado_job":   _str(fila, COL_CARGADO_JOB),
            "comentarios":   _str(fila, COL_COMENTARIOS),
        })
    return resultado


def load_jobs(excel_path: Path) -> dict[str, dict]:
    """
    Lee la hoja de Jobs ('Jobs', o 'Jobs FY26' si esa no existe).
    Retorna dict {proyecto_strip: {...}}.
    El engagement se normaliza con .strip().
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "Jobs" in wb.sheetnames:
        ws = wb["Jobs"]
    else:
        ws = wb["Jobs FY26"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    jobs = {}
    for fila in rows[1:]:
        if not fila[0]:
            continue
        proyecto = str(fila[0]).strip()
        eng_raw = fila[2]
        engagement = str(eng_raw).strip() if eng_raw else None
        # "None" textual puede aparecer si la celda tiene la cadena "None"
        if engagement in ("None", ""):
            engagement = None
        jobs[proyecto] = {
            "gerente":       fila[1],
            "engagement":    engagement,
            "eaf":           fila[3],
            "comentario_job":fila[4],
            "tipo_actividad":fila[5],
            "tipo":          fila[6],
            "cliente":       fila[7],
        }
    return jobs


def load_rates(excel_path: Path) -> dict[str, float]:
    """Lee 'Rates FY26'. Retorna dict {rank: multiplicador}."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Rates FY26"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    rates = {}
    for fila in rows[1:]:
        if fila[0]:
            rates[str(fila[0]).strip()] = fila[1]
    return rates


# ── helpers privados ──────────────────────────────────────────────────────────

def _str(fila: tuple, idx: int) -> str | None:
    """Extrae celda como string limpio, o None si está vacía."""
    if idx >= len(fila):
        return None
    val = fila[idx]
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _num(fila: tuple, idx: int) -> float:
    """Extrae celda como número. Retorna 0 si vacía o no numérica."""
    if idx >= len(fila):
        return 0.0
    val = fila[idx]
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0
