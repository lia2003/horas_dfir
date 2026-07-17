"""
Función 1 — Correos a gerentes (solicitud de horas).

Por cada proyecto Cliente de la semana (excluyendo LE y LE Gestión) genera:
  - <Proyecto>.txt  : wording del correo listo para enviar
  - <Proyecto>.xlsx : tabla con header negro / texto blanco
"""

from __future__ import annotations

from collections import defaultdict
from datetime import timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .excel_reader import (
    PROYECTOS_EXCLUIDOS,
    get_semana_rows,
    load_jobs,
    parse_semana_fecha,
)
from .utils import (
    guardar_excel,
    confirmar,
    fmt_fecha,
    fmt_fecha_corta,
    sanitizar_nombre_archivo,
)


def generar_correos(excel_path: Path, semana_nombre: str, output_dir: Path) -> None:
    lunes  = parse_semana_fecha(semana_nombre)
    viernes = lunes + timedelta(days=4)

    print(f"\n  Semana: {fmt_fecha(lunes)} - {fmt_fecha(viernes)}")

    filas = get_semana_rows(excel_path, semana_nombre)
    jobs  = load_jobs(excel_path)

    # ── 1. Filtrar filas Cliente válidas ──────────────────────────────────────
    filas_cliente: list[dict] = []
    sin_match: set[str] = set()

    for f in filas:
        if f["tipo"] != "Cliente":
            continue
        proy = (f["proyecto"] or "").strip()
        if not proy:
            continue
        if proy in PROYECTOS_EXCLUIDOS:
            continue
        if proy not in jobs:
            sin_match.add(proy)
            continue
        filas_cliente.append(f)

    # ── 2. Alertar proyectos sin match en Jobs FY26 ───────────────────────────
    if sin_match:
        print(f"\n  ATENCIÓN: Los siguientes proyectos no están en Jobs FY26:")
        for p in sorted(sin_match):
            print(f"    - {p}")
        print()
        nuevos_agregados = _ofrecer_agregar_jobs(sin_match, excel_path)
        if nuevos_agregados:
            jobs = load_jobs(excel_path)  # recargar tras escritura
            # Re-evaluar filas que antes no tenían match
            for f in filas:
                if f["tipo"] != "Cliente":
                    continue
                proy = (f["proyecto"] or "").strip()
                if proy in nuevos_agregados and proy in jobs:
                    filas_cliente.append(f)

    if not filas_cliente:
        print("  No hay filas Cliente para generar correos.")
        return

    # ── 3. Agrupar por proyecto ───────────────────────────────────────────────
    por_proyecto: dict[str, list[dict]] = defaultdict(list)
    for f in filas_cliente:
        por_proyecto[(f["proyecto"] or "").strip()].append(f)

    # ── 4. Generar archivos ───────────────────────────────────────────────────
    carpeta = output_dir / semana_nombre
    carpeta.mkdir(parents=True, exist_ok=True)

    generados: list[str] = []
    for proy, filas_proy in sorted(por_proyecto.items()):
        info      = jobs.get(proy, {})
        gerente   = info.get("gerente") or "Sin gerente"
        engagement = info.get("engagement")

        if not engagement:
            print(f"\n  ADVERTENCIA: '{proy}' no tiene engagement en Jobs FY26. "
                  "El correo se generará de todas formas.")

        _escribir_txt(proy, gerente, filas_proy, lunes, viernes, carpeta)
        _escribir_xlsx(proy, filas_proy, carpeta)
        generados.append(proy)

    print(f"\n  Correos generados ({len(generados)}):")
    for p in generados:
        print(f"    OK: {p}")
    print(f"\n  Carpeta de salida: {carpeta}")


# ── helpers ───────────────────────────────────────────────────────────────────

def _ofrecer_agregar_jobs(proyectos: set[str], excel_path: Path) -> set[str]:
    """Pregunta si el usuario quiere agregar proyectos faltantes a Jobs FY26."""
    agregados: set[str] = set()
    for proy in sorted(proyectos):
        if not confirmar(f"  ¿Agregar '{proy}' a Jobs FY26 ahora?"):
            print(f"    '{proy}' se omitirá de los correos.")
            continue

        gerente    = input(f"    Gerente de '{proy}': ").strip()
        engagement = input(f"    Engagement (Enter si no tiene aún): ").strip() or None
        tipo_act   = input(f"    Tipo de actividad (ej. DI-APOYO RD, Enter si no aplica): ").strip() or None

        print(f"\n    Resumen: Proyecto={proy} | Gerente={gerente} "
              f"| Engagement={engagement} | Tipo actividad={tipo_act}")
        if not confirmar("    ¿Confirmar?"):
            print(f"    '{proy}' se omitirá.")
            continue

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Jobs FY26"]
        ws.append([proy, gerente, engagement, 0, None, tipo_act, "Cliente", None])
        guardar_excel(wb, excel_path)
        wb.close()
        print(f"    '{proy}' agregado a Jobs FY26.")
        agregados.add(proy)

    return agregados


def _escribir_txt(
    proy: str,
    gerente: str,
    filas: list[dict],
    lunes,
    viernes,
    carpeta: Path,
) -> None:
    """Genera el .txt con asunto + cuerpo del correo."""
    asunto = f"Asistencia en proyecto {proy} - Semana {fmt_fecha_corta(lunes)}"

    # Tabla en texto plano para el cuerpo
    cabecera = f"{'Nombre':<25} {'Rank':<12} {'Proyecto':<20} {'Tarea':<55} {'Horas':>6}"
    separador = "-" * len(cabecera)
    filas_txt = [cabecera, separador]
    for f in filas:
        tarea = (f["tarea"] or "").strip()
        filas_txt.append(
            f"{(f['nombre'] or ''):<25} {(f['rank'] or ''):<12} "
            f"{proy:<20} {tarea:<55} {f['horas']:>6}"
        )

    tabla = "\n".join(filas_txt)

    gerente_corto = gerente.split()[0] if gerente else gerente

    cuerpo = (
        f"Hola {gerente_corto},\n\n"
        f"Le compartimos las horas incurridas en el proyecto {proy} "
        f"hasta hoy {fmt_fecha(viernes)}.\n\n"
        "Con el objetivo de poder contabilizar las horas y el ingreso generado "
        "por estas actividades, agradeceríamos nos pueda brindar el engagement "
        "donde cargar estas horas:\n\n"
        f"{tabla}\n\n"
        "Quedamos atentos a tu respuesta.\n\n"
        "Saludos,"
    )

    contenido = f"ASUNTO: {asunto}\n{'=' * 60}\n\n{cuerpo}"
    archivo = carpeta / f"{sanitizar_nombre_archivo(proy)}.txt"
    archivo.write_text(contenido, encoding="utf-8")


def _escribir_xlsx(proy: str, filas: list[dict], carpeta: Path) -> None:
    """Genera el .xlsx con la tabla (header negro / texto blanco)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = proy[:31]  # máximo 31 caracteres en nombre de hoja

    encabezados = ["Nombre", "Rank", "Proyecto", "Tarea", "Cantidad de horas"]
    anchos      = [25,       12,     20,          60,      20]

    fill_negro  = PatternFill("solid", fgColor="000000")
    font_blanco = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    font_normal = Font(name="Calibri", size=11)
    alin_centro = Alignment(horizontal="center", vertical="center")
    alin_izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    alin_der    = Alignment(horizontal="center",  vertical="center")

    # Fila de encabezado
    for col_idx, (h, ancho) in enumerate(zip(encabezados, anchos), start=1):
        celda = ws.cell(row=1, column=col_idx, value=h)
        celda.fill = fill_negro
        celda.font = font_blanco
        celda.alignment = alin_centro
        ws.column_dimensions[celda.column_letter].width = ancho

    ws.row_dimensions[1].height = 20

    # Filas de datos
    for fila_idx, f in enumerate(filas, start=2):
        tarea = (f["tarea"] or "").strip()
        datos = [f["nombre"], f["rank"], proy, tarea, f["horas"]]
        for col_idx, val in enumerate(datos, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=val)
            celda.font = font_normal
            celda.alignment = alin_izq if col_idx == 4 else alin_der

    archivo = carpeta / f"{sanitizar_nombre_archivo(proy)}.xlsx"
    wb.save(archivo)
