"""Generación de Reporte_RD.xlsx a partir de las horas de la Función 3.

Incluye TODOS los proyectos Cliente de la semana, tanto los que se mandan
en el mensaje al equipo (aprobados manualmente por gerente, columna
APROBADAS = ☑) como los que se marcaron como excluidos del mensaje
(no pasan por aprobación manual, columna APROBADAS = ☐ por defecto pero
editable en Excel).

Agrega (o reemplaza si ya existía) una hoja nombrada con la fecha del lunes
de la semana activa (DD-MM-YYYY) en el archivo indicado por 'reporte_rd_path'
en config.json, sin tocar las hojas de semanas anteriores.

Columnas (igual formato que el Reporte_RD ya existente):
  A: NSR Rate            E: APROBADAS
  B: Person Name         F: PROYECTO
  C: Engagement Number   G: EAF (de la hoja Jobs)
  D: Hours                H: SUBTOTAL = A * D * (1 + G)
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .utils import guardar_excel

HEADERS = [
    "NSR Rate", "Person Name", "Engagement Number", "Hours",
    "APROBADAS", "PROYECTO", "EAF", "SUBTOTAL",
]

RATE_ANDREA = 361

# Checkbox visual en la columna APROBADAS: se guarda como texto y se elige
# mediante dropdown de validación de datos (Excel no soporta insertar
# checkboxes reales de formulario vía openpyxl).
MARCA_SI = "☑"  # ☑
MARCA_NO = "☐"  # ☐

# Texto EXACTO de "Person Name" (rank + apellido, nombre) para cada persona
# conocida del equipo, fijo sin importar cómo esté escrito el rank o el
# nombre en la fila de la hoja semanal de ese proyecto (algunos proyectos
# registran a la misma persona con un rank/formato distinto, ej. "Incharge"
# en vez de "STAFF 2"). Se matchea buscando que AMBAS palabras clave (nombre
# y apellido) estén contenidas en el nombre crudo de la hoja semanal.
ROSTER_PERSON_NAME: list[tuple[tuple[str, str], str]] = [
    (("rodrigo", "delgado"),    "SENIOR MANAGER 1 - Delgado, Rodrigo Eugenio"),
    (("rodrigo", "beuzeville"), "SENIOR 2 - Beuzeville, Rodrigo Arturo"),
    (("marcelo", "rojas"),      "STAFF 2 - Rojas, Marcelo Jon"),
    (("alvaro", "barco"),       "STAFF 1 - Barco, Alvaro Joaquin"),
    (("manuel", "zambrano"),    "STAFF 1 - Zambrano, Manuel Nazaret"),
    (("marcelo", "carrion"),    "STAFF 1 - Carrion, Marcelo Andre"),
    (("andrea", "neira"),       "STAFF 1 - Neira, Andrea Valeria"),
    (("daniel", "cabrera"),     "INTERN (CS) - Cabrera, Daniel Sebastian"),
    (("lia", "arancibia"),      "INTERN (CS) - Arancibia, Lia Mariel"),
]


def _apellido_nombre(nombre: str) -> str:
    """'Nombre(s) Apellido' (como está en la hoja semanal) -> 'Apellido, Nombre(s)'.

    Fallback solo para personas que no están en ROSTER_PERSON_NAME.
    """
    partes = nombre.split()
    if len(partes) < 2:
        return nombre
    apellido = partes[-1]
    resto = " ".join(partes[:-1])
    return f"{apellido}, {resto}"


def _person_name_fijo(nombre: str) -> tuple[str, str] | None:
    """Busca a la persona en ROSTER_PERSON_NAME por nombre y apellido.

    Retorna (rank, texto_completo) si hay match, o None si no está en la lista.
    """
    n = (nombre or "").lower()
    for (clave1, clave2), texto in ROSTER_PERSON_NAME:
        if clave1 in n and clave2 in n:
            rank = texto.split(" - ", 1)[0]
            return rank, texto
    return None


def generar_reporte_rd(
    reporte_rd_path: Path,
    lunes: date,
    datos_aprobados: dict[str, dict],
    rank_por_nombre: dict[str, str],
    jobs: dict[str, dict],
    rates: dict[str, float],
    prorateos: dict[str, float],
    nombre_andrea: str | None,
    nombre_daniel: str | None,
) -> None:
    intern_rate = next((v for k, v in rates.items() if "intern" in k.lower()), 0)

    def _rate(nombre: str, rank: str) -> float:
        if nombre_daniel and nombre == nombre_daniel:
            v = intern_rate
        elif nombre_andrea and nombre == nombre_andrea:
            v = RATE_ANDREA
        else:
            v = rates.get(rank, 0)
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    filas: list[dict] = []
    for proy, datos in datos_aprobados.items():
        eng = datos["engagement"]
        aprobada = datos.get("aprobada", True)
        info_job = jobs.get(proy, {})
        try:
            eaf = float(info_job.get("eaf") or 0)
        except (ValueError, TypeError):
            eaf = 0.0
        for nombre, horas in datos["horas_aprobadas"].items():
            fijo = _person_name_fijo(nombre)
            if fijo:
                rank, person = fijo
            else:
                rank = rank_por_nombre.get(nombre, "")
                nombre_fmt = _apellido_nombre(nombre)
                person = f"{rank} - {nombre_fmt}" if rank else nombre_fmt

            factor = prorateos.get(nombre)
            if nombre_andrea and nombre == nombre_andrea and factor is not None:
                # Para Andrea el prorrateo NO se aplica a las Horas: se deja
                # como factor explícito dentro de la fórmula del SUBTOTAL.
                h = horas
                prorateo_formula = factor
            else:
                h = round(horas * factor, 1) if factor is not None else horas
                prorateo_formula = None
            filas.append({
                "rate":              _rate(nombre, rank),
                "person":            person,
                "engagement":        eng,
                "horas":             h,
                "proyecto":          proy,
                "eaf":               eaf,
                "aprobada":          aprobada,
                "prorateo_formula":  prorateo_formula,
            })

    if not filas:
        print("  No hay horas aprobadas para incluir en el Reporte_RD.")
        return

    filas.sort(key=lambda r: (-r["rate"], r["person"]))

    nombre_hoja = lunes.strftime("%d-%m-%Y")

    if reporte_rd_path.exists():
        wb = openpyxl.load_workbook(reporte_rd_path)
    else:
        reporte_rd_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

    if nombre_hoja in wb.sheetnames:
        wb.remove(wb[nombre_hoja])
    ws = wb.create_sheet(nombre_hoja)

    hdr_font   = Font(bold=True, color="FFFFFF")
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    center_aln = Alignment(horizontal="center", vertical="center")
    left_aln   = Alignment(horizontal="left", vertical="center")

    for col_idx, titulo in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center_aln

    for i, w in enumerate([10, 38, 20, 9, 11, 16, 8, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    fila_actual = 2
    for f in filas:
        ws.cell(row=fila_actual, column=1, value=f["rate"]).alignment = center_aln
        ws.cell(row=fila_actual, column=2, value=f["person"]).alignment = left_aln
        ws.cell(row=fila_actual, column=3, value=f["engagement"]).alignment = left_aln
        c_h = ws.cell(row=fila_actual, column=4, value=f["horas"])
        c_h.alignment = center_aln
        c_h.number_format = "0.0"
        marca = MARCA_SI if f["aprobada"] else MARCA_NO
        ws.cell(row=fila_actual, column=5, value=marca).alignment = center_aln
        ws.cell(row=fila_actual, column=6, value=f["proyecto"]).alignment = left_aln
        c_eaf = ws.cell(row=fila_actual, column=7, value=f["eaf"])
        c_eaf.alignment = center_aln
        c_eaf.number_format = "0.000"
        formula_sub = f"=A{fila_actual}*D{fila_actual}*(1+G{fila_actual})"
        if f["prorateo_formula"] is not None:
            formula_sub += f"*{f['prorateo_formula']}"
        c_sub = ws.cell(row=fila_actual, column=8, value=formula_sub)
        c_sub.alignment = center_aln
        c_sub.number_format = "0.00"
        fila_actual += 1

    if fila_actual > 2:
        dv = DataValidation(
            type="list",
            formula1=f'"{MARCA_SI},{MARCA_NO}"',
            allow_blank=False,
        )
        dv.error = f"Elige {MARCA_SI} (aprobado) o {MARCA_NO} (no aprobado)."
        dv.errorTitle = "Valor inválido"
        ws.add_data_validation(dv)
        dv.add(f"E2:E{fila_actual - 1}")

    fila_total = fila_actual + 1
    ws.cell(row=fila_total, column=7, value="TOTAL").font = Font(bold=True)
    c_total = ws.cell(
        row=fila_total, column=8,
        value=f"=SUM(H2:H{fila_actual - 1})",
    )
    c_total.font = Font(bold=True)
    c_total.number_format = "0.00"
    ws.cell(row=fila_total, column=9,  value="<--------")
    ws.cell(row=fila_total, column=10, value="EN CASO TODO ESTE APROBADO")

    guardar_excel(wb, reporte_rd_path)

    print(f"\n  Reporte_RD actualizado: hoja '{nombre_hoja}' -> {reporte_rd_path}")
